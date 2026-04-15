from __future__ import annotations

import json
import math
import os
import re
import threading
import time
from asyncio import to_thread
from collections import Counter
from pathlib import Path
from typing import Any

from src.backend.knowledge.evidence_organizer import source_family
from src.backend.knowledge.opendataloader_pdf import OpenDataLoaderPdfResult, build_pdf_documents_with_opendataloader
from src.backend.knowledge.types import Evidence, IndexStatus
from src.backend.runtime.config import get_settings
from pydantic import PrivateAttr


HEADING_PATTERN = re.compile(r"^(#{1,6})\s+(.*)$")
ALNUM_PATTERN = re.compile(r"[A-Za-z0-9_]+")
CHINESE_BLOCK_PATTERN = re.compile(r"[\u4e00-\u9fff]+")
SUPPORTED_FILE_SUFFIXES = {".md", ".json", ".txt", ".pdf", ".xlsx", ".xls"}
TEXT_FILE_ENCODINGS = ("utf-8", "utf-8-sig", "gb18030", "gbk")
PDF_OVERVIEW_SECTION_LIMIT = 6
PDF_OVERVIEW_TABLE_LIMIT = 4
PDF_OVERVIEW_FIGURE_LIMIT = 3
PDF_OVERVIEW_SNIPPET_LIMIT = 3


def _load_llama_index_components():
    from llama_index.core import (  # pylint: disable=import-outside-toplevel
        Document,
        Settings as LlamaSettings,
        StorageContext,
        VectorStoreIndex,
        load_index_from_storage,
    )
    from llama_index.core.base.embeddings.base import BaseEmbedding  # pylint: disable=import-outside-toplevel
    from llama_index.embeddings.openai import OpenAIEmbedding  # pylint: disable=import-outside-toplevel
    from llama_index.embeddings.huggingface import HuggingFaceEmbedding  # pylint: disable=import-outside-toplevel

    return (
        Document,
        LlamaSettings,
        StorageContext,
        VectorStoreIndex,
        load_index_from_storage,
        BaseEmbedding,
        OpenAIEmbedding,
        HuggingFaceEmbedding,
    )


def _build_compatible_openai_embedding(
    *,
    model_name: str,
    api_key: str,
    api_base: str,
):
    from openai import OpenAI  # pylint: disable=import-outside-toplevel

    _, _, _, _, _, BaseEmbedding, _, _ = _load_llama_index_components()

    class CompatibleOpenAIEmbedding(BaseEmbedding):
        model_name: str
        api_key: str
        api_base: str

        _client: OpenAI = PrivateAttr()

        def model_post_init(self, __context: Any) -> None:
            self._client = OpenAI(api_key=self.api_key, base_url=self.api_base)

        @classmethod
        def class_name(cls) -> str:
            return "CompatibleOpenAIEmbedding"

        def _embed(self, inputs: list[str]) -> list[list[float]]:
            response = self._client.embeddings.create(model=self.model_name, input=inputs)
            return [list(item.embedding) for item in response.data]

        def _get_query_embedding(self, query: str) -> list[float]:
            return self._embed([query])[0]

        async def _aget_query_embedding(self, query: str) -> list[float]:
            return await to_thread(self._get_query_embedding, query)

        def _get_text_embedding(self, text: str) -> list[float]:
            return self._embed([text])[0]

        def _get_text_embeddings(self, texts: list[str]) -> list[list[float]]:
            return self._embed(texts)

    return CompatibleOpenAIEmbedding(
        model_name=model_name,
        api_key=api_key,
        api_base=api_base,
    )


class KnowledgeIndexer:
    def __init__(self) -> None:
        self.base_dir: Path | None = None
        self._vector_index: Any | None = None
        self._documents: list[dict[str, Any]] = []
        self._build_errors: list[dict[str, str]] = []
        self._build_stats: dict[str, Any] = {}
        self._lock = threading.Lock()
        self._building = False
        self._last_built_at: float | None = None
        self._avg_doc_length = 0.0
        self._document_frequencies: Counter[str] = Counter()
        self._vector_ready = False
        self._bm25_ready = False
        self._last_vector_error: str | None = None

    def configure(self, base_dir: Path) -> None:
        self.base_dir = base_dir
        self._storage_dir.mkdir(parents=True, exist_ok=True)
        self._vector_dir.mkdir(parents=True, exist_ok=True)
        self._bm25_dir.mkdir(parents=True, exist_ok=True)
        self._derived_dir.mkdir(parents=True, exist_ok=True)
        self._vector_index = None
        self._vector_ready = False
        self._load_manifest()

    def _pdf_parser_backend(self) -> str:
        configured = os.getenv("PDF_PARSER_BACKEND", "opendataloader").strip().lower()
        if configured == "legacy":
            return "legacy"
        return "opendataloader"

    def warm_start(self) -> None:
        if self.base_dir is None:
            return
        with self._lock:
            self._load_manifest()
            if self._supports_embeddings() and any(self._vector_dir.glob("*")):
                self._load_vector_index()

    @property
    def _knowledge_dir(self) -> Path:
        if self.base_dir is None:
            raise RuntimeError("KnowledgeIndexer is not configured")
        return self.base_dir / "knowledge"

    @property
    def _storage_dir(self) -> Path:
        if self.base_dir is None:
            raise RuntimeError("KnowledgeIndexer is not configured")
        return self.base_dir / "storage" / "knowledge"

    @property
    def _manifest_path(self) -> Path:
        return self._storage_dir / "manifest.json"

    @property
    def _vector_dir(self) -> Path:
        return self._storage_dir / "vector"

    @property
    def _bm25_dir(self) -> Path:
        return self._storage_dir / "bm25"

    @property
    def _derived_dir(self) -> Path:
        return self._storage_dir / "derived"

    @property
    def _ingestion_errors_path(self) -> Path:
        return self._derived_dir / "ingestion_errors.json"

    def _supports_embeddings(self) -> bool:
        settings = get_settings()
        return settings.embedding_provider == "local" or bool(settings.embedding_api_key)

    def _build_embed_model(self):
        settings = get_settings()
        _, _, _, _, _, _, OpenAIEmbedding, HuggingFaceEmbedding = _load_llama_index_components()
        if settings.embedding_provider == "local":
            return HuggingFaceEmbedding(model_name=settings.embedding_model)
        if settings.embedding_provider == "bailian":
            return _build_compatible_openai_embedding(
                model_name=settings.embedding_model,
                api_key=str(settings.embedding_api_key or ""),
                api_base=settings.embedding_base_url,
            )
        return OpenAIEmbedding(
            api_key=settings.embedding_api_key,
            api_base=settings.embedding_base_url,
            model=settings.embedding_model,
        )

    def status(self) -> IndexStatus:
        return IndexStatus(
            ready=bool(self._documents) and (self._vector_ready or self._bm25_ready),
            building=self._building,
            last_built_at=self._last_built_at,
            indexed_files=len({item["source_path"] for item in self._documents}),
            vector_ready=self._vector_ready,
            bm25_ready=self._bm25_ready,
            vector_error=self._last_vector_error,
        )

    def is_building(self) -> bool:
        return self._building

    def build_stats(self) -> dict[str, Any]:
        return dict(self._build_stats)

    def rebuild_index(self, *, build_vector: bool = True) -> None:
        if self.base_dir is None:
            return

        with self._lock:
            self._building = True
            try:
                self._documents = self._build_documents()
                self._write_manifest()
                self._prepare_bm25_stats()
                if build_vector:
                    self._build_vector_index()
                else:
                    self._vector_index = None
                    self._vector_ready = False
                self._last_built_at = time.time()
            finally:
                self._building = False

    def _relative_path(self, path: Path) -> str:
        if self.base_dir is None:
            return str(path)
        return str(path.relative_to(self.base_dir)).replace("\\", "/")

    def _build_documents(self) -> list[dict[str, Any]]:
        if not self._knowledge_dir.exists():
            self._build_stats = {}
            return []

        documents: list[dict[str, Any]] = []
        self._build_errors = []
        self._build_stats = {
            "pdf_parser_backend": self._pdf_parser_backend(),
            "source_type_counts": {},
            "pdf_parser": {
                "parsed_pdf_count": 0,
                "failed_pdf_count": 0,
                "failure_reasons": {},
                "chunk_counts": {"text": 0, "table": 0, "figure_caption": 0},
                "avg_chunk_length": 0.0,
                "page_available_rate": 0.0,
                "bbox_available_rate": 0.0,
                "structure_modes": {"struct_tree": 0, "heuristic": 0, "unknown": 0},
            },
        }
        pdf_paths: list[Path] = []
        for path in sorted(self._knowledge_dir.rglob("*")):
            if not path.is_file():
                continue
            suffix = path.suffix.lower()
            if suffix not in SUPPORTED_FILE_SUFFIXES:
                continue
            try:
                if suffix == ".md":
                    documents.extend(self._split_markdown(path))
                elif suffix == ".json":
                    documents.extend(self._split_json(path))
                elif suffix == ".txt":
                    documents.extend(self._split_text(path))
                elif suffix == ".pdf":
                    pdf_paths.append(path)
                elif suffix == ".xlsx":
                    documents.extend(self._split_excel(path))
                elif suffix == ".xls":
                    self._record_build_error(path, "parse", "Legacy .xls ingestion is not supported; save the workbook as .xlsx.")
            except Exception as exc:
                self._record_build_error(path, "parse", str(exc))

        if pdf_paths:
            pdf_result = self._split_pdfs(pdf_paths)
            documents.extend(pdf_result.documents)
            self._build_errors.extend(pdf_result.errors)
            if pdf_result.stats:
                self._build_stats["pdf_parser"] = pdf_result.stats
            documents.extend(self._build_pdf_family_overviews(pdf_result.documents))

        source_type_counts = Counter(str(item.get("source_type", "")).strip().lower() for item in documents if item.get("source_type"))
        self._build_stats["source_type_counts"] = dict(sorted(source_type_counts.items()))
        self._write_ingestion_errors()
        return documents

    def _build_pdf_family_overviews(self, documents: list[dict[str, Any]]) -> list[dict[str, Any]]:
        grouped: dict[str, list[dict[str, Any]]] = {}
        for item in documents:
            if str(item.get("source_type", "")).strip().lower() != "pdf":
                continue
            family = source_family(str(item.get("source_path", "") or ""))
            if not family:
                continue
            grouped.setdefault(family, []).append(item)

        overviews: list[dict[str, Any]] = []
        for family, items in grouped.items():
            source_path = next(
                (
                    str(item.get("source_path", "") or "")
                    for item in items
                    if str(item.get("source_path", "") or "").lower().endswith(".pdf")
                ),
                family,
            )
            title = Path(source_path).stem
            pages = sorted({int(item.get("page")) for item in items if isinstance(item.get("page"), int)})
            section_titles = [
                section
                for section in dict.fromkeys(
                    str(item.get("section_title", "") or "").strip()
                    for item in items
                    if str(item.get("section_title", "") or "").strip()
                )
            ][:PDF_OVERVIEW_SECTION_LIMIT]
            table_titles = [
                locator
                for locator in dict.fromkeys(
                    str(item.get("locator", "") or "").strip()
                    for item in items
                    if str(item.get("chunk_type", "") or "") == "table" and str(item.get("locator", "") or "").strip()
                )
            ][:PDF_OVERVIEW_TABLE_LIMIT]
            figure_titles = [
                snippet
                for snippet in dict.fromkeys(
                    " ".join(str(item.get("snippet", "") or "").split())[:180]
                    for item in items
                    if str(item.get("chunk_type", "") or "") == "figure-caption" and str(item.get("snippet", "") or "").strip()
                )
            ][:PDF_OVERVIEW_FIGURE_LIMIT]
            snippet_samples = [
                snippet
                for snippet in dict.fromkeys(
                    " ".join(str(item.get("snippet", "") or "").split())[:180]
                    for item in items
                    if str(item.get("chunk_type", "") or "") in {"text", "text-group", "table"} and str(item.get("snippet", "") or "").strip()
                )
            ][:PDF_OVERVIEW_SNIPPET_LIMIT]

            period_hints = sorted(
                {
                    match.group(0)
                    for match in re.finditer(r"(20\d{2}|Q[1-4]|前三季度|第三季度|年初至报告期末|本报告期)", title, flags=re.IGNORECASE)
                }
            )
            entity_hints = [
                token
                for token in dict.fromkeys(re.findall(r"[\u4e00-\u9fff]{2,16}(?:集团|重工|动力|科技|汽车|股份有限公司|公司|报告)", title))
                if token
            ]
            overview_lines = [
                f"Source PDF: {source_path}",
                f"Source family: {family}",
                f"Title: {title}",
            ]
            if entity_hints:
                overview_lines.append("Entity hints: " + " | ".join(entity_hints[:4]))
            if period_hints:
                overview_lines.append("Period hints: " + " | ".join(period_hints[:6]))
            if pages:
                overview_lines.append("Pages: " + ", ".join(str(page) for page in pages[:8]))
            if section_titles:
                overview_lines.append("Sections: " + " | ".join(section_titles))
            if table_titles:
                overview_lines.append("Table titles: " + " | ".join(table_titles))
            if figure_titles:
                overview_lines.append("Figure captions: " + " | ".join(figure_titles))
            if snippet_samples:
                overview_lines.append("Key snippets: " + " | ".join(snippet_samples))

            overviews.append(
                {
                    "doc_id": f"{source_path}::family_overview",
                    "parent_id": f"{source_path}::family_overview",
                    "source_path": source_path,
                    "source_type": "pdf",
                    "locator": "family overview",
                    "text": "\n".join(overview_lines),
                    "parent_text": "\n".join(overview_lines),
                    "page": pages[0] if pages else None,
                    "bbox": None,
                    "element_type": "overview",
                    "section_title": title,
                    "chunk_type": "family_overview",
                }
            )
        return overviews

    def _split_pdfs(self, paths: list[Path]) -> OpenDataLoaderPdfResult:
        if self._pdf_parser_backend() == "legacy":
            documents: list[dict[str, Any]] = []
            errors: list[dict[str, str]] = []
            chunk_lengths: list[int] = []
            page_count = 0
            bbox_count = 0
            for path in paths:
                before_errors = len(self._build_errors)
                chunks = self._split_pdf_legacy(path)
                documents.extend(chunks)
                chunk_lengths.extend(len(str(item.get("text", "") or "")) for item in chunks)
                page_count += sum(1 for item in chunks if item.get("page") is not None)
                bbox_count += sum(1 for item in chunks if item.get("bbox") is not None)
                if len(self._build_errors) > before_errors:
                    errors.extend(self._build_errors[before_errors:])
            total_chunks = len(documents)
            stats = {
                "backend": "legacy",
                "parsed_pdf_count": len({str(item.get('source_path', '')) for item in documents}),
                "failed_pdf_count": len(errors),
                "failure_reasons": dict(Counter(item.get("message", "") for item in errors)),
                "chunk_counts": {"text": total_chunks, "table": 0, "figure_caption": 0},
                "avg_chunk_length": (sum(chunk_lengths) / len(chunk_lengths)) if chunk_lengths else 0.0,
                "page_available_rate": (page_count / total_chunks) if total_chunks else 0.0,
                "bbox_available_rate": (bbox_count / total_chunks) if total_chunks else 0.0,
                "structure_modes": {"struct_tree": 0, "heuristic": len(paths), "unknown": 0},
            }
            return OpenDataLoaderPdfResult(documents=documents, errors=[], stats=stats)

        try:
            return build_pdf_documents_with_opendataloader(
                base_dir=self.base_dir or Path("."),
                pdf_paths=paths,
                derived_root=self._derived_dir / "opendataloader",
            )
        except Exception as exc:
            message = str(exc).strip() or exc.__class__.__name__
            errors = [
                {
                    "source_path": self._relative_path(path),
                    "stage": "pdf_preflight",
                    "message": message,
                }
                for path in paths
            ]
            stats = {
                "backend": "opendataloader",
                "parsed_pdf_count": 0,
                "failed_pdf_count": len(paths),
                "failure_reasons": {message: len(paths)},
                "chunk_counts": {"text": 0, "table": 0, "figure_caption": 0},
                "avg_chunk_length": 0.0,
                "page_available_rate": 0.0,
                "bbox_available_rate": 0.0,
                "structure_modes": {"struct_tree": 0, "heuristic": 0, "unknown": len(paths)},
            }
            return OpenDataLoaderPdfResult(documents=[], errors=errors, stats=stats)

    def _split_pdf(self, path: Path) -> list[dict[str, Any]]:
        return self._split_pdfs([path]).documents

    def _record_build_error(self, path: Path, stage: str, message: str) -> None:
        self._build_errors.append(
            {
                "source_path": self._relative_path(path),
                "stage": stage,
                "message": message,
            }
        )

    def _write_ingestion_errors(self) -> None:
        self._ingestion_errors_path.write_text(
            json.dumps({"errors": self._build_errors, "stats": self._build_stats}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def _read_text_file(self, path: Path) -> str:
        for encoding in TEXT_FILE_ENCODINGS:
            try:
                return path.read_text(encoding=encoding)
            except UnicodeDecodeError:
                continue
        return path.read_bytes().decode("utf-8", errors="replace")

    def _make_text_chunks(
        self,
        *,
        source_path: str,
        source_type: str,
        parent_id: str,
        locator_prefix: str,
        text: str,
        parent_text: str | None = None,
        metadata: dict[str, Any] | None = None,
    ) -> list[dict[str, Any]]:
        cleaned_text = text.strip()
        if not cleaned_text:
            return []

        chunks: list[dict[str, Any]] = []
        paragraphs = [part.strip() for part in re.split(r"\n\s*\n+", cleaned_text) if part.strip()]
        if not paragraphs:
            paragraphs = [cleaned_text]

        for paragraph_index, paragraph in enumerate(paragraphs, start=1):
            slices = [paragraph[index : index + 1200] for index in range(0, len(paragraph), 1200)] or [paragraph]
            for slice_index, slice_text in enumerate(slices, start=1):
                locator = f"{locator_prefix} / 段落 {paragraph_index}"
                if len(slices) > 1:
                    locator = f"{locator}.{slice_index}"
                item = {
                    "doc_id": f"{parent_id}::child::{paragraph_index}.{slice_index}",
                    "parent_id": parent_id,
                    "source_path": source_path,
                    "source_type": source_type,
                    "locator": locator,
                    "text": slice_text,
                    "parent_text": (parent_text or cleaned_text).strip(),
                }
                if metadata:
                    item.update(metadata)
                chunks.append(item)
        return chunks

    def _split_markdown(self, path: Path) -> list[dict[str, Any]]:
        text = self._read_text_file(path)
        source_path = self._relative_path(path)
        sections: list[tuple[list[str], list[str]]] = []
        heading_stack: list[str] = []
        current_lines: list[str] = []

        def flush_section() -> None:
            if not current_lines:
                return
            heading_path = heading_stack[:] if heading_stack else [path.stem]
            sections.append((heading_path, current_lines[:]))

        for raw_line in text.splitlines():
            match = HEADING_PATTERN.match(raw_line)
            if not match:
                current_lines.append(raw_line)
                continue

            flush_section()
            current_lines = [raw_line]
            level = len(match.group(1))
            title = match.group(2).strip()
            heading_stack = heading_stack[: level - 1]
            heading_stack.append(title)

        flush_section()
        if not sections:
            sections = [([path.stem], text.splitlines())]

        chunks: list[dict[str, Any]] = []
        for section_index, (heading_path, lines) in enumerate(sections, start=1):
            section_text = "\n".join(lines).strip()
            if not section_text:
                continue
            parent_id = f"{source_path}::{' > '.join(heading_path)}"
            paragraphs = [part.strip() for part in re.split(r"\n\s*\n+", section_text) if part.strip()]
            if not paragraphs:
                paragraphs = [section_text]

            for paragraph_index, paragraph in enumerate(paragraphs, start=1):
                content = paragraph.strip()
                if not content:
                    continue
                slices = [content[index : index + 1200] for index in range(0, len(content), 1200)] or [content]
                for slice_index, slice_text in enumerate(slices, start=1):
                    locator = f"{' > '.join(heading_path)} / 段落 {paragraph_index}"
                    if len(slices) > 1:
                        locator = f"{locator}.{slice_index}"
                    chunks.append(
                        {
                            "doc_id": f"{parent_id}::child::{paragraph_index}.{slice_index}",
                            "parent_id": parent_id,
                            "source_path": source_path,
                            "source_type": "md",
                            "locator": locator,
                            "text": slice_text,
                            "parent_text": section_text,
                            "section_index": section_index,
                        }
                    )
        return chunks

    def _split_json(self, path: Path) -> list[dict[str, Any]]:
        source_path = self._relative_path(path)
        payload = json.loads(self._read_text_file(path))
        if not isinstance(payload, list):
            return []

        chunks: list[dict[str, Any]] = []
        for index, item in enumerate(payload, start=1):
            if not isinstance(item, dict):
                continue
            question = str(item.get("question", "")).strip()
            answer = str(item.get("answer", "")).strip()
            label = str(item.get("label", "")).strip()
            url = str(item.get("url", "")).strip()
            if not question and not answer:
                continue

            record_id = str(item.get("record_id") or item.get("id") or index)
            locator = f"记录 {record_id}"
            parts = []
            if question:
                parts.append(f"Question: {question}")
            if answer:
                parts.append(f"Answer: {answer}")
            if label:
                parts.append(f"Label: {label}")
            if url:
                parts.append(f"URL: {url}")
            text = "\n".join(parts)
            parent_id = f"{source_path}::record::{record_id}"
            chunks.append(
                {
                    "doc_id": f"{parent_id}::child::1",
                    "parent_id": parent_id,
                    "source_path": source_path,
                    "source_type": "json",
                    "locator": locator,
                    "text": text,
                    "parent_text": text,
                    "record_id": record_id,
                }
            )
        return chunks

    def _split_text(self, path: Path) -> list[dict[str, Any]]:
        source_path = self._relative_path(path)
        text = self._read_text_file(path).strip()
        if not text:
            return []
        return self._make_text_chunks(
            source_path=source_path,
            source_type="txt",
            parent_id=f"{source_path}::text",
            locator_prefix=path.stem,
            text=text,
            metadata={"file_type": "txt"},
        )

    def _split_pdf_legacy(self, path: Path) -> list[dict[str, Any]]:
        try:
            from pypdf import PdfReader  # pylint: disable=import-outside-toplevel
        except ImportError:
            from PyPDF2 import PdfReader  # pylint: disable=import-outside-toplevel

        source_path = self._relative_path(path)
        reader = PdfReader(str(path))
        chunks: list[dict[str, Any]] = []
        extracted_pages = 0
        total_pages = len(reader.pages)

        for page_number, page in enumerate(reader.pages, start=1):
            page_text = str(page.extract_text() or "").strip()
            if not page_text:
                continue
            extracted_pages += 1
            parent_id = f"{source_path}::page::{page_number}"
            chunks.extend(
                self._make_text_chunks(
                    source_path=source_path,
                    source_type="pdf",
                    parent_id=parent_id,
                    locator_prefix=f"页 {page_number}",
                    text=page_text,
                    metadata={
                        "file_type": "pdf",
                        "page": page_number,
                        "bbox": None,
                        "element_type": "page",
                        "chunk_type": "text",
                        "section_title": None,
                        "total_pages": total_pages,
                    },
                )
            )

        if not chunks:
            self._record_build_error(path, "extract", "No extractable text found in PDF.")
        elif extracted_pages < total_pages:
            self._record_build_error(
                path,
                "extract",
                f"Extracted text from {extracted_pages}/{total_pages} PDF pages; some pages returned no text.",
            )
        return chunks

    def _normalize_cell_value(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _split_excel(self, path: Path) -> list[dict[str, Any]]:
        from openpyxl import load_workbook  # pylint: disable=import-outside-toplevel

        source_path = self._relative_path(path)
        source_type = path.suffix.lower().lstrip(".")
        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
        chunks: list[dict[str, Any]] = []

        for sheet_index, sheet_name in enumerate(workbook.sheetnames, start=1):
            worksheet = workbook[sheet_name]
            non_empty_rows: list[tuple[int, list[str]]] = []
            for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                normalized_row = [self._normalize_cell_value(cell) for cell in row]
                if any(cell for cell in normalized_row):
                    non_empty_rows.append((row_number, normalized_row))

            if not non_empty_rows:
                continue

            header_row_number, header_row = non_empty_rows[0]
            headers = [value or f"Column {index}" for index, value in enumerate(header_row, start=1)]
            data_rows = non_empty_rows[1:]
            sheet_parent_id = f"{source_path}::sheet::{sheet_name}"

            overview_lines = [
                f"Sheet: {sheet_name}",
                f"Header row: {header_row_number}",
                f"Headers: {', '.join(headers)}",
                f"Data row count: {len(data_rows)}",
            ]
            for sample_row_number, sample_row in data_rows[:3]:
                row_pairs = [
                    f"{headers[index]}={value}"
                    for index, value in enumerate(sample_row)
                    if value and index < len(headers)
                ]
                if row_pairs:
                    overview_lines.append(f"Sample row {sample_row_number}: " + "; ".join(row_pairs))

            chunks.extend(
                self._make_text_chunks(
                    source_path=source_path,
                    source_type=source_type,
                    parent_id=f"{sheet_parent_id}::overview",
                    locator_prefix=f"Sheet {sheet_name} 概览",
                    text="\n".join(overview_lines),
                    metadata={
                        "file_type": source_type,
                        "sheet": sheet_name,
                        "sheet_index": sheet_index,
                        "row_start": header_row_number,
                        "row_end": header_row_number,
                    },
                )
            )

            for group_start in range(0, len(data_rows), 20):
                group = data_rows[group_start : group_start + 20]
                if not group:
                    continue
                row_start = group[0][0]
                row_end = group[-1][0]
                row_lines = [
                    f"Sheet: {sheet_name}",
                    f"Headers: {', '.join(headers)}",
                ]
                for row_number, row_values in group:
                    row_pairs = []
                    for column_index, value in enumerate(row_values):
                        if not value:
                            continue
                        header = headers[column_index] if column_index < len(headers) else f"Column {column_index + 1}"
                        row_pairs.append(f"{header}={value}")
                    if row_pairs:
                        row_lines.append(f"Row {row_number}: " + "; ".join(row_pairs))

                if len(row_lines) <= 2:
                    continue

                chunks.extend(
                    self._make_text_chunks(
                        source_path=source_path,
                        source_type=source_type,
                        parent_id=f"{sheet_parent_id}::rows::{row_start}-{row_end}",
                        locator_prefix=f"Sheet {sheet_name} / 行 {row_start}-{row_end}",
                        text="\n".join(row_lines),
                        metadata={
                            "file_type": source_type,
                            "sheet": sheet_name,
                            "sheet_index": sheet_index,
                            "row_start": row_start,
                            "row_end": row_end,
                        },
                    )
                )

        workbook.close()
        if not chunks:
            self._record_build_error(path, "extract", "No non-empty sheets or rows were extracted from the workbook.")
        return chunks

    def _write_manifest(self) -> None:
        payload = {
            "built_at": time.time(),
            "documents": self._documents,
            "errors": self._build_errors,
            "stats": self._build_stats,
        }
        self._manifest_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def _load_manifest(self) -> None:
        if not self._manifest_path.exists():
            self._documents = []
            self._build_stats = {}
            self._bm25_ready = False
            return
        try:
            payload = json.loads(self._manifest_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            self._documents = []
            self._build_errors = []
            self._build_stats = {}
            self._bm25_ready = False
            return
        self._documents = list(payload.get("documents", []))
        raw_errors = payload.get("errors", [])
        self._build_errors = raw_errors if isinstance(raw_errors, list) else []
        raw_stats = payload.get("stats", {})
        self._build_stats = raw_stats if isinstance(raw_stats, dict) else {}
        self._last_built_at = payload.get("built_at")
        self._vector_ready = False
        self._prepare_bm25_stats()

    def _prepare_bm25_stats(self) -> None:
        if not self._documents:
            self._avg_doc_length = 0.0
            self._document_frequencies = Counter()
            self._bm25_ready = False
            return

        self._document_frequencies = Counter()
        doc_lengths: list[int] = []
        for item in self._documents:
            tokens = self._tokenize(str(item.get("text", "")))
            item["tokens"] = tokens
            doc_lengths.append(len(tokens))
            for token in set(tokens):
                self._document_frequencies[token] += 1

        self._avg_doc_length = sum(doc_lengths) / max(1, len(doc_lengths))
        self._bm25_ready = True

    def _build_vector_index(self) -> None:
        if not self._supports_embeddings() or not self._documents:
            self._vector_index = None
            self._vector_ready = False
            return

        def vector_metadata(item: dict[str, Any]) -> dict[str, Any]:
            metadata: dict[str, Any] = {
                "source_path": str(item.get("source_path", "") or ""),
                "source_type": str(item.get("source_type", "") or ""),
                "parent_id": str(item.get("parent_id", "") or "")[:120] or None,
                "page": item.get("page"),
                "bbox": item.get("bbox"),
                "element_type": str(item.get("element_type", "") or "")[:80] or None,
                "section_title": str(item.get("section_title", "") or "")[:80] or None,
                "chunk_type": str(item.get("chunk_type", "") or "")[:40] or None,
            }
            return {key: value for key, value in metadata.items() if value not in (None, "")}

        try:
            Document, LlamaSettings, _, VectorStoreIndex, _, _, _, _ = _load_llama_index_components()
            LlamaSettings.embed_model = self._build_embed_model()
            documents = [
                Document(
                    text=str(item["text"]),
                    metadata=vector_metadata(item),
                )
                for item in self._documents
            ]
            self._vector_index = VectorStoreIndex.from_documents(documents)
            self._vector_index.storage_context.persist(persist_dir=str(self._vector_dir))
            self._vector_ready = True
            self._last_vector_error = None
        except Exception as exc:
            self._vector_index = None
            self._vector_ready = False
            self._last_vector_error = str(exc).strip() or exc.__class__.__name__

    def _load_vector_index(self) -> None:
        if not self._supports_embeddings():
            self._vector_index = None
            self._vector_ready = False
            self._last_vector_error = "Embeddings are not configured"
            return
        if not list(self._vector_dir.glob("*")):
            self._vector_index = None
            self._vector_ready = False
            self._last_vector_error = "Persisted vector directory does not exist"
            return
        try:
            _, LlamaSettings, StorageContext, _, load_index_from_storage, _, _, _ = _load_llama_index_components()
            LlamaSettings.embed_model = self._build_embed_model()
            storage_context = StorageContext.from_defaults(persist_dir=str(self._vector_dir))
            self._vector_index = load_index_from_storage(storage_context)
            self._vector_ready = True
            self._last_vector_error = None
        except Exception as exc:
            self._vector_index = None
            self._vector_ready = False
            self._last_vector_error = str(exc).strip() or exc.__class__.__name__

    def _ensure_loaded(self) -> None:
        if not self._documents:
            self._load_manifest()

    def _matches_path_filters(self, source_path: str, path_filters: list[str] | None) -> bool:
        if not path_filters:
            return True
        normalized = source_path.replace("\\", "/")
        for path_filter in path_filters:
            candidate = path_filter.replace("\\", "/").strip()
            if not candidate:
                continue
            if normalized == candidate or normalized.startswith(f"{candidate}/"):
                return True
        return False

    def _matches_chunk_type_filters(self, chunk_type: str | None, chunk_types: list[str] | None) -> bool:
        if not chunk_types:
            return True
        normalized = str(chunk_type or "").strip().lower()
        allowed = {str(item).strip().lower() for item in chunk_types if str(item).strip()}
        return normalized in allowed

    def retrieve_vector(
        self,
        query: str,
        *,
        top_k: int = 4,
        path_filters: list[str] | None = None,
        chunk_types: list[str] | None = None,
    ) -> list[Evidence]:
        self._ensure_loaded()
        if self._vector_index is None and self._supports_embeddings() and any(self._vector_dir.glob("*")):
            self._load_vector_index()
        if self._vector_index is None:
            return []

        retriever = self._vector_index.as_retriever(similarity_top_k=max(top_k * 4, top_k))
        try:
            results = retriever.retrieve(query)
        except Exception:
            return []

        payload: list[Evidence] = []
        for item in results:
            node = getattr(item, "node", item)
            metadata = getattr(node, "metadata", {}) or {}
            source_path = str(metadata.get("source_path", ""))
            if not self._matches_path_filters(source_path, path_filters):
                continue
            chunk_type = str(metadata.get("chunk_type", "") or "") or None
            if not self._matches_chunk_type_filters(chunk_type, chunk_types):
                continue
            text = getattr(node, "text", "") or getattr(node, "get_content", lambda: "")()
            raw_parent_id = metadata.get("parent_id")
            parent_id = str(raw_parent_id).strip() if raw_parent_id else None
            locator = str(metadata.get("locator", "") or "").strip()
            if not locator:
                page = metadata.get("page")
                section_title = str(metadata.get("section_title", "") or "").strip()
                element_type = str(metadata.get("element_type", "") or "").strip() or "evidence"
                locator_parts: list[str] = []
                if page is not None:
                    locator_parts.append(f"页 {page}")
                if section_title:
                    locator_parts.append(section_title)
                locator_parts.append(element_type)
                locator = " / ".join(locator_parts)
            payload.append(
                Evidence(
                    source_path=source_path,
                    source_type=str(metadata.get("source_type", "unknown")),
                    locator=locator,
                    snippet=str(text).strip(),
                    channel="vector",
                    score=float(getattr(item, "score", 0.0) or 0.0),
                    parent_id=parent_id,
                    page=metadata.get("page"),
                    bbox=metadata.get("bbox"),
                    element_type=str(metadata.get("element_type", "") or "") or None,
                    section_title=str(metadata.get("section_title", "") or "") or None,
                    derived_json_path=str(metadata.get("derived_json_path", "") or "") or None,
                    derived_markdown_path=str(metadata.get("derived_markdown_path", "") or "") or None,
                    chunk_type=chunk_type,
                )
            )
            if len(payload) >= top_k:
                break
        return payload

    def retrieve_bm25(
        self,
        query: str,
        *,
        top_k: int = 4,
        path_filters: list[str] | None = None,
        query_hints: list[str] | None = None,
        chunk_types: list[str] | None = None,
    ) -> list[Evidence]:
        self._ensure_loaded()
        if not self._documents or not self._bm25_ready:
            return []

        hints = " ".join(query_hints or [])
        query_tokens = self._tokenize(f"{query} {hints}".strip())
        if not query_tokens:
            return []

        candidates = [
            item
            for item in self._documents
            if self._matches_path_filters(str(item["source_path"]), path_filters)
            and self._matches_chunk_type_filters(str(item.get("chunk_type", "") or "") or None, chunk_types)
        ]
        if not candidates:
            candidates = list(self._documents)

        scores: list[tuple[dict[str, Any], float]] = []
        corpus_size = max(1, len(self._documents))
        k1 = 1.5
        b = 0.75
        for item in candidates:
            doc_tokens = item.get("tokens", [])
            if not doc_tokens:
                continue
            token_counts = Counter(doc_tokens)
            doc_len = len(doc_tokens)
            score = 0.0
            for token in query_tokens:
                if token not in token_counts:
                    continue
                df = self._document_frequencies.get(token, 0)
                if df <= 0:
                    continue
                idf = math.log(1 + ((corpus_size - df + 0.5) / (df + 0.5)))
                freq = token_counts[token]
                denominator = freq + k1 * (1 - b + b * (doc_len / max(1.0, self._avg_doc_length)))
                score += idf * ((freq * (k1 + 1)) / max(denominator, 1e-9))
            if score > 0:
                scores.append((item, score))

        scores.sort(key=lambda item: item[1], reverse=True)
        payload: list[Evidence] = []
        for item, score in scores[:top_k]:
            raw_parent_id = item.get("parent_id")
            parent_id = str(raw_parent_id).strip() if raw_parent_id else None
            payload.append(
                Evidence(
                    source_path=str(item["source_path"]),
                    source_type=str(item["source_type"]),
                    locator=str(item["locator"]),
                    snippet=str(item["text"]).strip(),
                    channel="bm25",
                    score=score,
                    parent_id=parent_id,
                    page=item.get("page"),
                    bbox=item.get("bbox"),
                    element_type=str(item.get("element_type", "") or "") or None,
                    section_title=str(item.get("section_title", "") or "") or None,
                    derived_json_path=str(item.get("derived_json_path", "") or "") or None,
                    derived_markdown_path=str(item.get("derived_markdown_path", "") or "") or None,
                    chunk_type=str(item.get("chunk_type", "") or "") or None,
                )
            )
        return payload

    def _tokenize(self, text: str) -> list[str]:
        lowered = text.lower()
        tokens: list[str] = []
        tokens.extend(ALNUM_PATTERN.findall(lowered))
        for match in CHINESE_BLOCK_PATTERN.findall(lowered):
            tokens.extend(list(match))
            if len(match) > 1:
                tokens.extend(match[index : index + 2] for index in range(len(match) - 1))
        return tokens


knowledge_indexer = KnowledgeIndexer()
